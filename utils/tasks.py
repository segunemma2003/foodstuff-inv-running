"""
Celery task definitions.

All tasks are imported by the worker via `include=["utils.tasks"]` in celery_app.py.
The API side only calls `.delay()` / `.apply_async()` — it never runs the body directly,
so the API response is always dispatched in < 1 ms.

Generated files (PDFs, reports) are stored in S3 under the `jobs/` prefix.
Uploaded input files are stored in S3 under the `uploads/` prefix and deleted after use.

Task categories
───────────────
  Email        → send_email_task
  PDF          → generate_quotation_pdf_task, generate_invoice_pdf_task
  Reports      → generate_report_task
  Bulk upload  → process_cost_price_bulk_task, process_product_bulk_task
"""
from celery_app import celery_app
import os
from sqlalchemy import func
from utils.email import send_email
import base64

INVOICE_PRIMARY_RECIPIENT = "foodstuffstoreinvoices@gmail.com"


# ─── Email ────────────────────────────────────────────────────────────────────

@celery_app.task(bind=True, name="send_email", max_retries=3, default_retry_delay=60)
def send_email_task(self, to: str, subject: str, html: str, text: str = ""):
    """Send a single email; retries up to 3 times on SMTP failure."""
    try:
        send_email(to, subject, html, text)
    except Exception as exc:
        raise self.retry(exc=exc)


@celery_app.task(bind=True, name="send_email_with_attachment", max_retries=3, default_retry_delay=60)
def send_email_with_attachment_task(
    self,
    recipients: list[str],
    subject: str,
    html: str,
    text: str,
    filename: str,
    mime_type: str,
    content_b64: str,
):
    """Send one attachment email to multiple recipients."""
    try:
        normalized = list(dict.fromkeys([e.strip() for e in recipients if e and e.strip()]))
        if not normalized:
            raise ValueError("No recipients provided")
        content = base64.b64decode(content_b64.encode("utf-8"))
        outcomes: list[dict[str, str | None]] = []
        for recipient in normalized:
            try:
                send_email(
                    to=recipient,
                    subject=subject,
                    html=html,
                    text=text,
                    attachments=[(filename, content, mime_type)],
                )
                outcomes.append({"recipient": recipient, "status": "delivered", "error": None})
            except Exception as send_exc:
                outcomes.append({"recipient": recipient, "status": "failed", "error": str(send_exc)})
        if not any(item["status"] == "delivered" for item in outcomes):
            raise ValueError("All recipients failed")
        return {"delivery_outcomes": outcomes}
    except Exception as exc:
        raise self.retry(exc=exc)


# ─── PDF generation ──────────────────────────────────────────────────────────

@celery_app.task(bind=True, name="generate_quotation_pdf")
def generate_quotation_pdf_task(self, quotation_id: int):
    from database import SessionLocal
    import models
    from utils.pdf_generator import generate_quotation_pdf
    from utils.s3 import upload_bytes

    db = SessionLocal()
    try:
        q = db.query(models.Quotation).filter(models.Quotation.id == quotation_id).first()
        if q is None:
            raise ValueError(f"Quotation {quotation_id} not found")

        pdf_bytes = generate_quotation_pdf(q)
        s3_key = f"jobs/{self.request.id}.pdf"
        upload_bytes(s3_key, pdf_bytes, "application/pdf")
        return {
            "s3_key": s3_key,
            "filename": f"{q.quotation_number}.pdf",
            "content_type": "application/pdf",
        }
    finally:
        db.close()


@celery_app.task(bind=True, name="generate_invoice_pdf")
def generate_invoice_pdf_task(self, invoice_id: int):
    from database import SessionLocal
    import models
    from utils.pdf_generator import generate_invoice_pdf
    from utils.s3 import upload_bytes

    db = SessionLocal()
    try:
        inv = db.query(models.Invoice).filter(models.Invoice.id == invoice_id).first()
        if inv is None:
            raise ValueError(f"Invoice {invoice_id} not found")

        pdf_bytes = generate_invoice_pdf(inv)
        s3_key = f"jobs/{self.request.id}.pdf"
        upload_bytes(s3_key, pdf_bytes, "application/pdf")
        return {
            "s3_key": s3_key,
            "filename": f"{inv.invoice_number}.pdf",
            "content_type": "application/pdf",
        }
    finally:
        db.close()


# ─── Report generation ───────────────────────────────────────────────────────

@celery_app.task(bind=True, name="generate_report")
def generate_report_task(self, report_type: str, params: dict):
    """
    Build an Excel report and upload it to S3.
    `params` is a plain dict with string values (JSON-safe).
    """
    from io import BytesIO
    from database import SessionLocal
    from utils.report_builder import build_report
    from utils.s3 import upload_bytes

    db = SessionLocal()
    try:
        wb, filename = build_report(report_type, params, db)
        buf = BytesIO()
        wb.save(buf)
        content_type = (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        s3_key = f"jobs/{self.request.id}.xlsx"
        upload_bytes(s3_key, buf.getvalue(), content_type)
        return {
            "s3_key": s3_key,
            "filename": filename,
            "content_type": content_type,
        }
    finally:
        db.close()


# ─── Send quotation to customer ──────────────────────────────────────────────

@celery_app.task(bind=True, name="send_quotation_to_customer", max_retries=3, default_retry_delay=60)
def send_quotation_to_customer_task(self, quotation_id: int):
    """Generate the quotation PDF and email it to the customer with the PDF attached."""
    from database import SessionLocal
    import models
    from utils.pdf_generator import generate_quotation_pdf
    from utils.email import send_email, tpl_quotation_to_customer

    db = SessionLocal()
    try:
        q = db.query(models.Quotation).filter(models.Quotation.id == quotation_id).first()
        if q is None:
            raise ValueError(f"Quotation {quotation_id} not found")

        customer_email = q.customer.email if q.customer else None
        if not customer_email:
            raise ValueError(f"Customer has no email address")

        pdf_bytes = generate_quotation_pdf(q)
        subject, html, text = tpl_quotation_to_customer(
            q.quotation_number,
            q.customer.customer_name,
            float(q.total_amount),
        )
        recipients = list(dict.fromkeys([customer_email, INVOICE_PRIMARY_RECIPIENT]))
        for recipient in recipients:
            send_email(
                to=recipient,
                subject=subject,
                html=html,
                text=text,
                attachments=[(f"{q.quotation_number}.pdf", pdf_bytes, "application/pdf")],
            )
    except Exception as exc:
        raise self.retry(exc=exc)
    finally:
        db.close()


@celery_app.task(bind=True, name="send_invoice_to_recipients", max_retries=3, default_retry_delay=60)
def send_invoice_to_recipients_task(self, invoice_id: int, recipients: list[str]):
    """Generate/download invoice PDF and send to recipient list in background."""
    from database import SessionLocal
    import models
    from utils.pdf_generator import generate_invoice_pdf
    from utils.email import tpl_invoice_to_customer, SMTP_USER, SMTP_PASSWORD
    from utils.s3 import download_bytes

    db = SessionLocal()
    try:
        inv = db.query(models.Invoice).filter(models.Invoice.id == invoice_id).first()
        if inv is None:
            raise ValueError(f"Invoice {invoice_id} not found")
        if not SMTP_USER or not SMTP_PASSWORD:
            raise ValueError("SMTP is not configured (SMTP_USER/SMTP_PASSWORD missing)")

        normalized_recipients = list(dict.fromkeys([e.strip() for e in recipients if e and e.strip()]))
        if not normalized_recipients:
            raise ValueError("No recipients provided")

        if inv.custom_pdf_s3_key:
            pdf_bytes = download_bytes(inv.custom_pdf_s3_key)
        else:
            bank_accounts = (
                db.query(models.PaymentAccount)
                .filter(models.PaymentAccount.is_active == True)
                .order_by(models.PaymentAccount.is_default.desc())
                .all()
            )
            pdf_bytes = generate_invoice_pdf(inv, bank_accounts=bank_accounts)

        customer_name = inv.customer.customer_name if inv.customer else "Customer"
        subject, html, text = tpl_invoice_to_customer(
            invoice_number=inv.invoice_number,
            customer_name=customer_name,
            total=float(inv.total_amount),
        )
        outcomes: list[dict[str, str | None]] = []
        for recipient in normalized_recipients:
            try:
                send_email(
                    to=recipient,
                    subject=subject,
                    html=html,
                    text=text,
                    attachments=[(f"{inv.invoice_number}.pdf", pdf_bytes, "application/pdf")],
                )
                outcomes.append({"recipient": recipient, "status": "delivered", "error": None})
            except Exception as send_exc:
                outcomes.append({"recipient": recipient, "status": "failed", "error": str(send_exc)})
        if not any(item["status"] == "delivered" for item in outcomes):
            raise ValueError("All recipients failed")
        return {"delivery_outcomes": outcomes}
    except Exception as exc:
        raise self.retry(exc=exc)
    finally:
        db.close()


# ─── Bulk uploads ─────────────────────────────────────────────────────────────
# Global celery_app defaults are 90s / 120s (celery_app.py). Excel imports can exceed
# that for large files or slow DB; these tasks override with longer limits.

@celery_app.task(
    bind=True,
    name="process_cost_price_bulk",
    soft_time_limit=900,
    time_limit=960,
)
def process_cost_price_bulk_task(self, s3_key: str, user_id: int):
    """
    Download an Excel file from S3, parse cost price records, insert into DB.
    The S3 object is deleted after processing (success or failure).
    """
    from io import BytesIO
    from database import SessionLocal
    import models
    import openpyxl
    from datetime import date as date_type
    from utils.s3 import download_bytes, delete_object

    db = SessionLocal()
    try:
        raw = download_bytes(s3_key)
        wb = openpyxl.load_workbook(BytesIO(raw))
        ws = wb.active
        headers = [
            str(c.value).strip().lower() if c.value else ""
            for c in next(ws.iter_rows(min_row=1, max_row=1))
        ]
        created, errors = 0, []

        def _norm(val):
            if val is None:
                return ""
            return " ".join(str(val).strip().lower().split())

        def _norm_key(val):
            import re
            return re.sub(r"[^a-z0-9]+", "", _norm(val))

        def _parse_cost_price_cell(val):
            """Accept numbers, Excel-formatted strings (1,000 $, ₦1,000, NGN 500, etc.)."""
            from decimal import Decimal, InvalidOperation
            import re
            if val is None or val == "":
                return None
            if isinstance(val, bool):
                return None
            if isinstance(val, (int, float)):
                try:
                    return Decimal(str(val))
                except InvalidOperation:
                    return None
            s = str(val).strip()
            if not s:
                return None
            s = s.replace("\u00a0", " ").replace(",", "")
            for token in (
                "NGN", "ngn", "N ", "$", "USD", "usd", "₦", "£", "€",
            ):
                s = s.replace(token, "")
            s = s.strip()
            m = re.search(r"-?\d+(?:\.\d+)?", s)
            if not m:
                return None
            try:
                d = Decimal(m.group(0))
                return d if d >= 0 else None
            except InvalidOperation:
                return None

        def _generate_unique_sku(product_name: str, resolved_market_id: int | None) -> str:
            import re
            base_name = re.sub(r"[^A-Z0-9]+", "-", str(product_name or "").upper()).strip("-") or "PRODUCT"
            base_name = base_name[:24]
            market_part = f"M{resolved_market_id}" if resolved_market_id else "M0"
            counter = 1
            while True:
                candidate = f"{base_name}-{market_part}-{counter:03d}"
                exists = db.query(models.Product).filter(models.Product.sku == candidate).first()
                if not exists:
                    return candidate
                counter += 1

        def _cell_as_text(val):
            if val is None:
                return ""
            if isinstance(val, str):
                return val.strip()
            if isinstance(val, bool):
                return ""
            return str(val).strip()

        def _cost_cell_blank(val):
            if val is None:
                return True
            if isinstance(val, str) and not val.strip():
                return True
            return False

        consecutive_blank_rows = 0
        BLANK_ROW_STOP = 3

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            data = dict(zip(headers, row))
            product_name = (data.get("product_name") or "").strip() if isinstance(data.get("product_name"), str) else data.get("product_name")
            market_name = (data.get("market_name") or "").strip() if isinstance(data.get("market_name"), str) else data.get("market_name")
            uom = data.get("uom") or data.get("unit_of_measure")
            uom = (uom or "").strip() if isinstance(uom, str) else uom
            cost_price = data.get("cost_price")

            pn_t = _cell_as_text(product_name)
            mn_t = _cell_as_text(market_name)
            uom_t = _cell_as_text(uom)

            if not pn_t and not mn_t and not uom_t and _cost_cell_blank(cost_price):
                consecutive_blank_rows += 1
                if consecutive_blank_rows >= BLANK_ROW_STOP:
                    break
                continue
            consecutive_blank_rows = 0

            if cost_price is None or not product_name or not market_name or not uom:
                continue

            parsed_price = _parse_cost_price_cell(cost_price)
            if parsed_price is None:
                errors.append(f"Row {row_num}: invalid cost_price {cost_price!r} — use a plain number (commas / ₦ / $ are OK)")
                continue

            market = db.query(models.ProductCategory).filter(
                func.lower(func.trim(models.ProductCategory.name)) == _norm(market_name)
            ).first()
            if not market:
                errors.append(f"Row {row_num}: market '{market_name}' not found")
                continue

            # Prefer strict match: product + market + uom
            product = db.query(models.Product).filter(
                func.lower(func.trim(models.Product.product_name)) == _norm(product_name),
                models.Product.category_id == market.id,
                func.lower(func.trim(func.coalesce(models.Product.unit_of_measure, ""))) == _norm(uom),
            ).first()

            # Fallback: if UOM does not match exactly, resolve by product+market
            # when this identifies a single product record.
            if not product:
                by_name = db.query(models.Product).filter(
                    func.lower(func.trim(models.Product.product_name)) == _norm(product_name),
                    models.Product.category_id == market.id,
                ).all()
                if not by_name:
                    # Last-resort tolerant name match (ignores spaces/punctuation like "50kg" vs "50 kg")
                    candidates = db.query(models.Product).filter(
                        models.Product.category_id == market.id
                    ).all()
                    target_key = _norm_key(product_name)
                    by_name = [p for p in candidates if _norm_key(p.product_name) == target_key]
                if len(by_name) == 1:
                    product = by_name[0]
                elif len(by_name) > 1:
                    available_uoms = sorted({(p.unit_of_measure or "").strip() or "—" for p in by_name})
                    # If multiple products share name and none matched exactly by uom,
                    # create a new product record using the provided uom.
                    pass

            if not product:
                product = models.Product(
                    product_name=str(product_name).strip(),
                    sku=_generate_unique_sku(str(product_name).strip(), market.id),
                    unit_of_measure=str(uom).strip() if uom is not None else None,
                    category_id=market.id,
                )
                db.add(product)
                db.flush()

            if product.category_id != market.id:
                errors.append(
                    f"Row {row_num}: product is linked to a different market than '{market_name}' — skipped (market not changed)"
                )
                continue

            db.add(models.CostPrice(
                product_id=product.id,
                cost_price=parsed_price,
                effective_date=date_type.today(),
                notes=None,
                created_by=user_id,
            ))
            created += 1

        db.commit()
        return {"created": created, "errors": errors}
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()
        delete_object(s3_key)


@celery_app.task(
    bind=True,
    name="process_product_bulk",
    soft_time_limit=900,
    time_limit=960,
)
def process_product_bulk_task(self, s3_key: str, user_id: int, market_id: int | None = None):
    """Download an Excel file from S3, parse and insert product records."""
    from io import BytesIO
    from database import SessionLocal
    import models
    import openpyxl
    from utils.s3 import download_bytes, delete_object

    db = SessionLocal()
    try:
        raw = download_bytes(s3_key)
        wb = openpyxl.load_workbook(BytesIO(raw))
        ws = wb.active
        headers = [
            str(c.value).strip().lower() if c.value else ""
            for c in next(ws.iter_rows(min_row=1, max_row=1))
        ]
        created, updated, errors = 0, 0, []

        def _generate_unique_sku(product_name: str, resolved_market_id: int | None) -> str:
            import re
            base_name = re.sub(r"[^A-Z0-9]+", "-", str(product_name or "").upper()).strip("-") or "PRODUCT"
            base_name = base_name[:24]
            market_part = f"M{resolved_market_id}" if resolved_market_id else "M0"
            counter = 1
            while True:
                candidate = f"{base_name}-{market_part}-{counter:03d}"
                exists = db.query(models.Product).filter(models.Product.sku == candidate).first()
                if not exists:
                    return candidate
                counter += 1

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            data = dict(zip(headers, row))
            name = (data.get("product_name") or "")
            if isinstance(name, str):
                name = name.strip()
            if not name:
                continue
            sku = data.get("sku") or None
            if isinstance(sku, str):
                sku = sku.strip() or None
            if sku and db.query(models.Product).filter(models.Product.sku == sku).first():
                errors.append(f"Row {row_num}: SKU '{sku}' already exists")
                continue
            market_name = (
                data.get("market_name")
                or data.get("market")
                or data.get("category_name")
                or data.get("category")
            )
            if isinstance(market_name, str):
                market_name = market_name.strip()
            resolved_market_id = market_id
            if resolved_market_id is None and market_name:
                cat = db.query(models.ProductCategory).filter(
                    func.lower(models.ProductCategory.name) == str(market_name).strip().lower()
                ).first()
                if cat:
                    resolved_market_id = cat.id
            if resolved_market_id is None:
                errors.append(f"Row {row_num}: market is required and must exist")
                continue
            if not sku:
                sku = _generate_unique_sku(name, resolved_market_id)
            unit_of_measure = data.get("unit_of_measure") or data.get("uom")
            if isinstance(unit_of_measure, str):
                unit_of_measure = unit_of_measure.strip() or None
            existing = db.query(models.Product).filter(
                models.Product.category_id == resolved_market_id,
                func.lower(models.Product.product_name) == str(name).strip().lower(),
            ).first()
            if existing:
                if sku:
                    existing.sku = sku
                if unit_of_measure:
                    existing.unit_of_measure = unit_of_measure
                updated += 1
            else:
                db.add(models.Product(
                    product_name=name,
                    sku=sku,
                    unit_of_measure=unit_of_measure,
                    category_id=resolved_market_id,
                ))
                created += 1

        db.commit()
        return {"created": created, "updated": updated, "errors": errors}
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()
        delete_object(s3_key)


@celery_app.task(
    bind=True,
    name="process_invoice_bulk_task",
    soft_time_limit=900,
    time_limit=960,
)
def process_invoice_bulk_task(self, s3_key: str, user_id: int):
    """
    Parse an Excel invoice import file and create invoices + line items.

    Rows sharing the same invoice_number (or same customer_name + invoice_date
    when invoice_number is blank) are grouped into one invoice.

    Required columns : customer_name, invoice_date, product_name, qty, unit_price
    Optional columns : invoice_number, due_date, payment_term, delivery_type, notes
    """
    from io import BytesIO
    from datetime import datetime, date as date_type
    from collections import OrderedDict
    from database import SessionLocal
    import models
    import openpyxl
    from utils.s3 import download_bytes, delete_object
    from utils.number_gen import next_invoice_number

    db = SessionLocal()
    try:
        raw = download_bytes(s3_key)
        wb = openpyxl.load_workbook(BytesIO(raw), data_only=True)
        ws = wb.active
        headers = [
            str(c.value).strip().lower().replace(" ", "_") if c.value else ""
            for c in next(ws.iter_rows(min_row=1, max_row=1))
        ]

        def _parse_date(val):
            if val is None:
                return None
            if isinstance(val, datetime):
                return val.date()
            if isinstance(val, date_type):
                return val
            try:
                return date_type.fromisoformat(str(val).strip())
            except ValueError:
                return None

        def _str(val, default=""):
            return str(val).strip() if val is not None else default

        groups: OrderedDict = OrderedDict()
        errors = []

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            data = dict(zip(headers, row))
            if not any(v for v in data.values()):
                continue

            customer_name = _str(data.get("customer_name"))
            invoice_date  = _parse_date(data.get("invoice_date"))
            product_name  = _str(data.get("product_name"))

            if not customer_name or not invoice_date or not product_name:
                errors.append(f"Row {row_num}: customer_name, invoice_date and product_name are required")
                continue

            try:
                qty        = float(data.get("qty") or data.get("quantity") or 0)
                unit_price = float(data.get("unit_price") or 0)
            except (ValueError, TypeError):
                errors.append(f"Row {row_num}: qty and unit_price must be numbers")
                continue

            if qty <= 0:
                errors.append(f"Row {row_num}: qty must be > 0")
                continue

            explicit_num = _str(data.get("invoice_number"))
            group_key    = explicit_num or f"{customer_name}|{invoice_date.isoformat()}"

            if group_key not in groups:
                due = _parse_date(data.get("due_date"))
                pt  = _str(data.get("payment_term"), "cash").lower().replace(" ", "_")
                dt  = _str(data.get("delivery_type"), "pickup").lower()
                groups[group_key] = {
                    "customer_name":   customer_name,
                    "invoice_date":    invoice_date,
                    "due_date":        due,
                    "payment_term":    pt,
                    "delivery_type":   dt,
                    "notes":           _str(data.get("notes")) or None,
                    "explicit_number": explicit_num or None,
                    "first_row":       row_num,
                    "items":           [],
                }

            groups[group_key]["items"].append({
                "product_name": product_name,
                "qty":          qty,
                "unit_price":   unit_price,
                "row_num":      row_num,
            })

        created = skipped = 0

        for group_key, grp in groups.items():
            customer = (
                db.query(models.Customer)
                .filter(models.Customer.customer_name.ilike(grp["customer_name"]))
                .first()
            )
            if not customer:
                errors.append(f"Row {grp['first_row']}: customer '{grp['customer_name']}' not found")
                skipped += 1
                continue

            try:
                delivery_type = models.DeliveryType(grp["delivery_type"])
            except ValueError:
                delivery_type = models.DeliveryType.pickup

            line_items = []
            invoice_total = 0.0

            for item_data in grp["items"]:
                product = (
                    db.query(models.Product)
                    .filter(models.Product.product_name.ilike(item_data["product_name"]))
                    .first()
                )
                if not product:
                    errors.append(f"Row {item_data['row_num']}: product '{item_data['product_name']}' not found")
                    continue

                latest_cost = (
                    db.query(models.CostPrice)
                    .filter(models.CostPrice.product_id == product.id)
                    .order_by(models.CostPrice.effective_date.desc())
                    .first()
                )
                cost_price    = float(latest_cost.cost_price) if latest_cost else 0.0
                line_total    = item_data["qty"] * item_data["unit_price"]
                invoice_total += line_total

                line_items.append(models.InvoiceItem(
                    product_id=product.id,
                    quantity=item_data["qty"],
                    uom=product.unit_of_measure,
                    cost_price=cost_price,
                    supply_markup_pct=0,
                    supply_markup_amount=0,
                    delivery_markup_pct=0,
                    delivery_markup_amount=0,
                    payment_term_markup_pct=0,
                    payment_term_markup_amount=0,
                    unit_price=item_data["unit_price"],
                    line_total=line_total,
                ))

            if not line_items:
                skipped += 1
                continue

            inv_number = grp["explicit_number"] or next_invoice_number(db)
            if db.query(models.Invoice).filter(models.Invoice.invoice_number == inv_number).first():
                errors.append(f"Invoice number '{inv_number}' already exists — skipped")
                skipped += 1
                continue

            db.add(models.Invoice(
                invoice_number=inv_number,
                quotation_id=None,
                customer_id=customer.id,
                invoice_date=grp["invoice_date"],
                due_date=grp["due_date"],
                payment_term=grp["payment_term"],
                delivery_type=delivery_type,
                notes=grp["notes"],
                total_amount=invoice_total,
                amount_paid=0,
                created_by=user_id,
                items=line_items,
            ))
            db.flush()
            created += 1

        db.commit()
        return {"created": created, "skipped": skipped, "errors": errors}
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()
        delete_object(s3_key)
