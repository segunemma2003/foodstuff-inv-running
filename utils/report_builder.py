"""
Pure workbook-building functions — no HTTP, no request context.
Used by both the Celery tasks (background) and was previously
used directly in the routers.

Each public function receives (db, params_dict) and returns (Workbook, filename).
params_dict values are all plain strings / None (JSON-safe for Celery).
"""
from datetime import date, datetime
from typing import Optional, Tuple

from sqlalchemy.orm import Session
from sqlalchemy import func, String
import openpyxl
from openpyxl.styles import Font

import models


def _not_cancelled_invoice_filter():
    return func.lower(models.Invoice.status.cast(String)) != "cancelled"


def _parse_date(val) -> Optional[date]:
    if val is None:
        return None
    if isinstance(val, date):
        return val
    return date.fromisoformat(str(val))


def _bold_row(ws, row_num: int = 1):
    for cell in ws[row_num]:
        cell.font = Font(bold=True)


# ─── Individual report builders ──────────────────────────────────────────────

def build_sales(db: Session, params: dict) -> Tuple[openpyxl.Workbook, str]:
    date_from = _parse_date(params.get("date_from"))
    date_to = _parse_date(params.get("date_to"))
    customer_id = params.get("customer_id")
    payment_term = params.get("payment_term")
    delivery_type = params.get("delivery_type")

    q = db.query(models.Invoice).filter(_not_cancelled_invoice_filter())
    if date_from:
        q = q.filter(models.Invoice.invoice_date >= date_from)
    if date_to:
        q = q.filter(models.Invoice.invoice_date <= date_to)
    if customer_id:
        q = q.filter(models.Invoice.customer_id == int(customer_id))
    if payment_term:
        q = q.filter(models.Invoice.payment_term == payment_term)
    if delivery_type:
        q = q.filter(models.Invoice.delivery_type == delivery_type)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Report"
    ws.append(["Invoice #", "Date", "Customer", "Payment Term", "Delivery", "Total", "Created By"])
    _bold_row(ws)
    for inv in q.order_by(models.Invoice.invoice_date).all():
        ws.append([
            inv.invoice_number, str(inv.invoice_date),
            inv.customer.customer_name if inv.customer else "",
            inv.payment_term, inv.delivery_type.value,
            float(inv.total_amount),
            inv.creator.full_name if inv.creator else "",
        ])
    return wb, "sales_report.xlsx"


def build_invoices(db: Session, params: dict) -> Tuple[openpyxl.Workbook, str]:
    date_from = _parse_date(params.get("date_from"))
    date_to = _parse_date(params.get("date_to"))
    q = db.query(models.Invoice)
    if date_from:
        q = q.filter(models.Invoice.invoice_date >= date_from)
    if date_to:
        q = q.filter(models.Invoice.invoice_date <= date_to)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoices"
    ws.append(["Invoice #", "Quotation #", "Date", "Customer", "Payment Term",
               "Delivery", "Due Date", "Total", "Status", "Created By"])
    _bold_row(ws)
    for inv in q.all():
        ws.append([
            inv.invoice_number,
            inv.quotation.quotation_number if inv.quotation else "",
            str(inv.invoice_date),
            inv.customer.customer_name if inv.customer else "",
            inv.payment_term, inv.delivery_type.value,
            str(inv.due_date) if inv.due_date else "",
            float(inv.total_amount), inv.status.value,
            inv.creator.full_name if inv.creator else "",
        ])
    return wb, "invoice_report.xlsx"


def build_quotations(db: Session, params: dict) -> Tuple[openpyxl.Workbook, str]:
    date_from = _parse_date(params.get("date_from"))
    date_to = _parse_date(params.get("date_to"))
    status = params.get("status")
    q = db.query(models.Quotation)
    if date_from:
        q = q.filter(models.Quotation.quotation_date >= date_from)
    if date_to:
        q = q.filter(models.Quotation.quotation_date <= date_to)
    if status:
        q = q.filter(models.Quotation.status == status)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Quotations"
    ws.append(["Quotation #", "Date", "Customer", "Payment Term",
               "Delivery", "Total", "Status", "Created By", "Approved By"])
    _bold_row(ws)
    for quot in q.all():
        ws.append([
            quot.quotation_number, str(quot.quotation_date),
            quot.customer.customer_name if quot.customer else "",
            quot.payment_term, quot.delivery_type.value,
            float(quot.total_amount), quot.status.value,
            quot.creator.full_name if quot.creator else "",
            quot.approver.full_name if quot.approver else "",
        ])
    return wb, "quotation_report.xlsx"


def build_customer_sales(db: Session, params: dict) -> Tuple[openpyxl.Workbook, str]:
    date_from = _parse_date(params.get("date_from"))
    date_to = _parse_date(params.get("date_to"))
    q = (
        db.query(
            models.Customer.id, models.Customer.customer_name,
            models.Customer.business_name, models.Customer.city,
            func.count(models.Invoice.id).label("invoice_count"),
            func.sum(models.Invoice.total_amount).label("total_sales"),
        )
        .join(models.Invoice, models.Invoice.customer_id == models.Customer.id)
        .filter(_not_cancelled_invoice_filter())
    )
    if date_from:
        q = q.filter(models.Invoice.invoice_date >= date_from)
    if date_to:
        q = q.filter(models.Invoice.invoice_date <= date_to)
    rows = q.group_by(models.Customer.id, models.Customer.customer_name,
                      models.Customer.business_name, models.Customer.city).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customer Sales"
    ws.append(["Customer ID", "Customer Name", "Business Name", "City", "# Invoices", "Total Sales"])
    _bold_row(ws)
    for r in rows:
        ws.append([r.id, r.customer_name, r.business_name or "",
                   r.city or "", r.invoice_count, float(r.total_sales or 0)])
    return wb, "customer_sales_report.xlsx"


def build_product_sales(db: Session, params: dict) -> Tuple[openpyxl.Workbook, str]:
    date_from = _parse_date(params.get("date_from"))
    date_to = _parse_date(params.get("date_to"))
    q = (
        db.query(
            models.Product.id, models.Product.product_name, models.Product.sku,
            func.sum(models.InvoiceItem.quantity).label("total_qty"),
            func.sum(models.InvoiceItem.line_total).label("total_revenue"),
            func.count(func.distinct(models.Invoice.id)).label("invoice_count"),
            func.count(func.distinct(models.Invoice.customer_id)).label("customer_count"),
        )
        .join(models.InvoiceItem, models.InvoiceItem.product_id == models.Product.id)
        .join(models.Invoice, models.Invoice.id == models.InvoiceItem.invoice_id)
        .filter(_not_cancelled_invoice_filter())
    )
    if date_from:
        q = q.filter(models.Invoice.invoice_date >= date_from)
    if date_to:
        q = q.filter(models.Invoice.invoice_date <= date_to)
    rows = q.group_by(models.Product.id, models.Product.product_name, models.Product.sku).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Product Sales"
    ws.append(["Product ID", "Product Name", "SKU", "Total Qty",
               "Total Revenue", "# Invoices", "# Customers"])
    _bold_row(ws)
    for r in rows:
        ws.append([r.id, r.product_name, r.sku or "", float(r.total_qty or 0),
                   float(r.total_revenue or 0), r.invoice_count, r.customer_count])
    return wb, "product_sales_report.xlsx"


def build_cost_price_history(db: Session, params: dict) -> Tuple[openpyxl.Workbook, str]:
    date_from = _parse_date(params.get("date_from"))
    date_to = _parse_date(params.get("date_to"))
    q = db.query(models.CostPrice)
    if date_from:
        q = q.filter(models.CostPrice.effective_date >= date_from)
    if date_to:
        q = q.filter(models.CostPrice.effective_date <= date_to)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cost Price History"
    ws.append(["Product ID", "Product Name", "SKU", "Cost Price",
               "Effective Date", "Notes", "Created By", "Created At"])
    _bold_row(ws)
    for cp in q.order_by(models.CostPrice.effective_date.desc()).all():
        ws.append([
            cp.product_id,
            cp.product.product_name if cp.product else "",
            cp.product.sku if cp.product else "",
            float(cp.cost_price), str(cp.effective_date),
            cp.notes or "",
            cp.creator.full_name if cp.creator else "",
            str(cp.created_at),
        ])
    return wb, "cost_price_history.xlsx"


def build_staff_performance(db: Session, params: dict) -> Tuple[openpyxl.Workbook, str]:
    date_from = _parse_date(params.get("date_from"))
    date_to = _parse_date(params.get("date_to"))
    users = db.query(models.User).filter(models.User.is_active == True).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Staff Performance"
    ws.append(["User ID", "Full Name", "Username", "Role",
               "Quotations Created", "Invoices Created", "Total Sales", "Conversion Rate %"])
    _bold_row(ws)
    for u in users:
        q_filter = [models.Quotation.created_by == u.id]
        i_filter = [models.Invoice.created_by == u.id, _not_cancelled_invoice_filter()]
        if date_from:
            q_filter.append(models.Quotation.quotation_date >= date_from)
            i_filter.append(models.Invoice.invoice_date >= date_from)
        if date_to:
            q_filter.append(models.Quotation.quotation_date <= date_to)
            i_filter.append(models.Invoice.invoice_date <= date_to)
        qc = db.query(func.count(models.Quotation.id)).filter(*q_filter).scalar() or 0
        ic_data = db.query(func.count(models.Invoice.id),
                           func.sum(models.Invoice.total_amount)).filter(*i_filter).first()
        ic = ic_data[0] or 0
        it = float(ic_data[1] or 0)
        conv = round(ic / qc * 100, 2) if qc else 0
        ws.append([u.id, u.full_name, u.username, u.role.value, qc, ic, it, conv])
    return wb, "staff_performance_report.xlsx"


# ─── Dispatcher ──────────────────────────────────────────────────────────────

_BUILDERS = {
    "sales": build_sales,
    "invoices": build_invoices,
    "quotations": build_quotations,
    "customer_sales": build_customer_sales,
    "product_sales": build_product_sales,
    "cost_price_history": build_cost_price_history,
    "staff_performance": build_staff_performance,
}


def build_report(report_type: str, params: dict, db: Session) -> Tuple[openpyxl.Workbook, str]:
    builder = _BUILDERS.get(report_type)
    if not builder:
        raise ValueError(f"Unknown report type: {report_type!r}. "
                         f"Valid: {list(_BUILDERS)}")
    return builder(db, params)
