from typing import List, Optional
from datetime import date

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from pydantic import BaseModel
from sqlalchemy.orm import Session
from sqlalchemy import func

from database import get_db
from dependencies import get_current_user, require_admin_or_manager, require_not_analyst, require_admin
import models
import schemas
from utils import audit
from utils.tasks import generate_invoice_pdf_task, send_invoice_to_recipients_task
from utils.make_integration import send_document_to_make_from_s3
from utils.queue_events import log_queue_event


class InvoiceSendEmailRequest(BaseModel):
    additional_emails: Optional[List[str]] = None

router = APIRouter(prefix="/invoices", tags=["Invoices"])
INVOICE_PRIMARY_RECIPIENT = "foodstuffstoreinvoices@gmail.com"


@router.get("/approved-quotations", response_model=List[schemas.QuotationOut])
def list_convertible_quotations(
    customer_id: Optional[int] = None,
    db: Session = Depends(get_db),
    _: models.User = Depends(get_current_user),
):
    """Return approved quotations that have not yet been converted to invoices."""
    convertible_quotation_query = (
        db.query(models.Quotation)
        .outerjoin(models.Invoice, models.Invoice.quotation_id == models.Quotation.id)
        .filter(
            models.Quotation.status == models.QuotationStatus.approved,
            models.Invoice.id == None,
        )
    )
    if customer_id:
        convertible_quotation_query = convertible_quotation_query.filter(models.Quotation.customer_id == customer_id)
    return convertible_quotation_query.order_by(models.Quotation.approved_at.desc()).all()


@router.get("/template")
def download_invoice_template():
    """Return a formatted Excel template for invoice import."""
    from io import BytesIO
    from datetime import date
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoice Import"

    headers = [
        ("invoice_number", 20),
        ("customer_name",  28),
        ("invoice_date",   16),
        ("due_date",       16),
        ("payment_term",   18),
        ("delivery_type",  16),
        ("product_name",   30),
        ("qty",            10),
        ("unit_price",     14),
        ("notes",          30),
    ]

    green_fill  = PatternFill("solid", fgColor="1E8449")
    alt_fill    = PatternFill("solid", fgColor="EAF4EE")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    body_font   = Font(size=10)
    thin_side   = Side(style="thin", color="CCCCCC")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # Header row
    for col, (name, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font      = header_font
        cell.fill      = green_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = thin_border
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 22

    # Sample data — two invoices, first has 2 line items
    today = date.today()
    sample = [
        # invoice_number  customer_name   inv_date         due_date          pay_term    delivery   product_name    qty   price    notes
        ("INV-2026-0001", "GENESIS GROUP", date(2026,1,15), date(2026,2,14),  "net_30",   "delivery","Rice 50kg",    10,   85000,  ""),
        ("INV-2026-0001", "GENESIS GROUP", date(2026,1,15), date(2026,2,14),  "net_30",   "delivery","Beans 50kg",    5,   45000,  ""),
        ("",              "ACME Corp",     date(2026,1,16), "",               "cash",     "pickup",  "Rice 50kg",     2,   85000,  "Urgent order"),
    ]

    date_fmt = "YYYY-MM-DD"
    for r, row in enumerate(sample, 2):
        fill = alt_fill if r % 2 == 0 else None
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font   = body_font
            cell.border = thin_border
            if fill:
                cell.fill = fill
            if isinstance(val, date):
                cell.number_format = date_fmt
                cell.alignment = Alignment(horizontal="center")
            elif c in (8, 9):  # qty, unit_price
                cell.alignment = Alignment(horizontal="right")
                if c == 9:
                    cell.number_format = "#,##0"

    ws.freeze_panes = "A2"  # freeze header row

    # Notes sheet
    ns = wb.create_sheet("How to Fill")
    ns.column_dimensions["A"].width = 18
    ns.column_dimensions["B"].width = 85
    ns["A1"] = "Column"
    ns["B1"] = "Instructions"
    ns["A1"].font = Font(bold=True, color="FFFFFF")
    ns["B1"].font = Font(bold=True, color="FFFFFF")
    ns["A1"].fill = green_fill
    ns["B1"].fill = green_fill

    instructions = [
        ("invoice_number", "Optional. Leave blank to auto-generate. Use the SAME value on multiple rows to group them into one invoice."),
        ("customer_name",  "REQUIRED. Must match a customer name in the system exactly (case-insensitive)."),
        ("invoice_date",   "REQUIRED. Date format: YYYY-MM-DD  (e.g. 2026-01-15)"),
        ("due_date",       "Optional. Date format: YYYY-MM-DD"),
        ("payment_term",   "Optional. Allowed values: cash  immediate  net_7  net_14  net_30  net_45  net_60  net_90.  Default: cash"),
        ("delivery_type",  "Optional. Allowed values: delivery  pickup.  Default: pickup"),
        ("product_name",   "REQUIRED. Must match a product name in the system exactly (case-insensitive)."),
        ("qty",            "REQUIRED. Quantity sold. Must be greater than 0."),
        ("unit_price",     "REQUIRED. Selling price per unit — numbers only, no currency symbol (e.g. 85000)."),
        ("notes",          "Optional. Any internal note for this invoice."),
    ]
    for i, (col, desc) in enumerate(instructions, 2):
        a = ns.cell(row=i, column=1, value=col)
        b = ns.cell(row=i, column=2, value=desc)
        a.font = Font(bold=True, size=10)
        b.font = Font(size=10)
        if i % 2 == 0:
            a.fill = alt_fill
            b.fill = alt_fill

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="invoice_import_template.xlsx"'},
    )


@router.post("/bulk-upload", response_model=schemas.JobEnqueuedResponse, status_code=202)
async def bulk_upload_invoices(
    file: UploadFile = File(...),
    current_user: models.User = Depends(require_not_analyst),
):
    """Upload an Excel file to import invoices. Returns a task_id — poll /api/v1/jobs/{task_id}."""
    import uuid
    from utils.s3 import upload_bytes
    from utils.tasks import process_invoice_bulk_task

    content = await file.read()
    s3_key = f"uploads/invoices_{uuid.uuid4()}.xlsx"
    upload_bytes(s3_key, content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    task = process_invoice_bulk_task.delay(s3_key, current_user.id)
    return schemas.JobEnqueuedResponse(
        task_id=task.id,
        message=f"Invoice import queued. Poll /api/v1/jobs/{task.id} for result.",
    )


@router.get("", response_model=List[schemas.InvoiceOut])
def list_invoices(
    skip: int = 0,
    limit: int = 50,
    customer_id: Optional[int] = None,
    status: Optional[str] = None,
    created_by: Optional[int] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    payment_term: Optional[str] = None,
    delivery_type: Optional[str] = None,
    db: Session = Depends(get_db),
    _: models.User = Depends(get_current_user),
):
    invoice_query = db.query(models.Invoice)
    if customer_id:
        invoice_query = invoice_query.filter(models.Invoice.customer_id == customer_id)
    if status:
        invoice_query = invoice_query.filter(models.Invoice.status == status)
    if created_by:
        invoice_query = invoice_query.filter(models.Invoice.created_by == created_by)
    if date_from:
        invoice_query = invoice_query.filter(models.Invoice.invoice_date >= date_from)
    if date_to:
        invoice_query = invoice_query.filter(models.Invoice.invoice_date <= date_to)
    if payment_term:
        invoice_query = invoice_query.filter(models.Invoice.payment_term == payment_term)
    if delivery_type:
        invoice_query = invoice_query.filter(models.Invoice.delivery_type == delivery_type)
    return invoice_query.order_by(models.Invoice.created_at.desc()).offset(skip).limit(limit).all()


def _delete_invoice_record(
    db: Session,
    invoice_id: int,
    current_user: models.User,
) -> schemas.MessageResponse:
    from utils.s3 import delete_object

    inv = db.query(models.Invoice).filter(models.Invoice.id == invoice_id).first()
    if not inv:
        raise HTTPException(404, "Invoice not found")

    quotation_id = inv.quotation_id
    inv_no = inv.invoice_number

    for payment in list(inv.payments):
        db.delete(payment)

    if inv.custom_pdf_s3_key:
        try:
            delete_object(inv.custom_pdf_s3_key)
        except Exception:
            pass

    audit.log(db, models.AuditAction.delete, models.AuditEntity.invoice, inv.id,
               current_user.id, description=f"Deleted invoice {inv_no}")
    db.delete(inv)

    if quotation_id:
        qt = db.query(models.Quotation).filter(models.Quotation.id == quotation_id).first()
        if qt and qt.status == models.QuotationStatus.converted:
            qt.status = models.QuotationStatus.approved

    return schemas.MessageResponse(message=f"Invoice {inv_no} deleted")


@router.post("/bulk-delete", response_model=schemas.BulkDeleteResult)
def bulk_delete_invoices(
    body: schemas.BulkIdsRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
):
    result = schemas.BulkDeleteResult()
    for iid in body.ids:
        try:
            _delete_invoice_record(db, iid, current_user)
            result.deleted += 1
        except HTTPException as he:
            result.failed.append({"id": iid, "detail": he.detail if isinstance(he.detail, str) else str(he.detail)})
    db.commit()
    return result


@router.post("", response_model=schemas.InvoiceOut, status_code=201)
def create_invoice(
    body: schemas.InvoiceCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_not_analyst),
):
    """Create an invoice directly (without a quotation)."""
    from utils.number_gen import next_invoice_number

    customer = db.query(models.Customer).filter(models.Customer.id == body.customer_id).first()
    if not customer:
        raise HTTPException(404, "Customer not found")

    if not body.items:
        raise HTTPException(400, "At least one line item is required")

    try:
        delivery_type = models.DeliveryType(body.delivery_type)
    except ValueError:
        raise HTTPException(400, f"Invalid delivery_type '{body.delivery_type}'")

    line_items = []
    total_amount = 0.0

    for item in body.items:
        product = db.query(models.Product).filter(models.Product.id == item.product_id).first()
        if not product:
            raise HTTPException(404, f"Product {item.product_id} not found")
        if not product.is_active:
            raise HTTPException(400, f"Product '{product.product_name}' is deactivated and cannot be used")

        latest_cost = (
            db.query(models.CostPrice)
            .filter(models.CostPrice.product_id == product.id)
            .order_by(models.CostPrice.effective_date.desc())
            .first()
        )
        cost_price = float(latest_cost.cost_price) if latest_cost else 0.0
        unit_price = float(item.unit_price)
        quantity   = float(item.quantity)
        line_total = unit_price * quantity
        total_amount += line_total

        line_items.append(models.InvoiceItem(
            product_id=product.id,
            quantity=quantity,
            uom=item.uom or product.unit_of_measure,
            cost_price=cost_price,
            supply_markup_pct=0,
            supply_markup_amount=0,
            delivery_markup_pct=0,
            delivery_markup_amount=0,
            payment_term_markup_pct=0,
            payment_term_markup_amount=0,
            unit_price=unit_price,
            line_total=line_total,
        ))

    inv = models.Invoice(
        invoice_number=next_invoice_number(db),
        quotation_id=None,
        customer_id=customer.id,
        invoice_date=body.invoice_date,
        due_date=body.due_date,
        payment_term=body.payment_term,
        delivery_type=delivery_type,
        notes=body.notes,
        total_amount=total_amount,
        amount_paid=0,
        created_by=current_user.id,
        items=line_items,
    )
    db.add(inv)
    db.commit()
    db.refresh(inv)
    audit.log(db, models.AuditAction.create, models.AuditEntity.invoice, inv.id,
              current_user.id, description=f"Created invoice {inv.invoice_number} directly")
    return inv


@router.get("/{invoice_id}", response_model=schemas.InvoiceOut)
def get_invoice(
    invoice_id: int,
    db: Session = Depends(get_db),
    _: models.User = Depends(get_current_user),
):
    inv = db.query(models.Invoice).filter(models.Invoice.id == invoice_id).first()
    if not inv:
        raise HTTPException(404, "Invoice not found")
    return inv


@router.get("/{invoice_id}/pdf")
def download_invoice_pdf(
    invoice_id: int,
    db: Session = Depends(get_db),
    _: models.User = Depends(get_current_user),
):
    """Stream invoice PDF. Serves uploaded PDF if present, otherwise generates from template."""
    from io import BytesIO
    from fastapi.responses import StreamingResponse

    inv = db.query(models.Invoice).filter(models.Invoice.id == invoice_id).first()
    if not inv:
        raise HTTPException(404, "Invoice not found")

    # Serve the uploaded PDF if one exists
    if inv.custom_pdf_s3_key:
        from utils.s3 import download_bytes
        pdf_bytes = download_bytes(inv.custom_pdf_s3_key)
        return StreamingResponse(
            BytesIO(pdf_bytes),
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{inv.invoice_number}.pdf"'},
        )

    from utils.pdf_generator import generate_invoice_pdf as gen_pdf

    bank_accounts = (
        db.query(models.PaymentAccount)
        .filter(models.PaymentAccount.is_active == True)
        .order_by(models.PaymentAccount.is_default.desc())
        .all()
    )
    paystack_payment = (
        db.query(models.Payment)
        .filter(
            models.Payment.invoice_id == invoice_id,
            models.Payment.paystack_payment_url.isnot(None),
            models.Payment.status == models.PaymentStatus.pending,
        )
        .order_by(models.Payment.created_at.desc())
        .first()
    )
    paystack_url = paystack_payment.paystack_payment_url if paystack_payment else None
    pdf_bytes = gen_pdf(inv, bank_accounts=bank_accounts, paystack_url=paystack_url)
    return StreamingResponse(
        BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{inv.invoice_number}.pdf"'},
    )


@router.get("/{invoice_id}/signed-pdf")
def download_signed_invoice_pdf(
    invoice_id: int,
    db: Session = Depends(get_db),
    _: models.User = Depends(get_current_user),
):
    """Stream only the uploaded signed invoice PDF."""
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    from utils.s3 import download_bytes

    inv = db.query(models.Invoice).filter(models.Invoice.id == invoice_id).first()
    if not inv:
        raise HTTPException(404, "Invoice not found")
    if not inv.custom_pdf_s3_key:
        raise HTTPException(404, "Signed invoice not uploaded yet")

    pdf_bytes = download_bytes(inv.custom_pdf_s3_key)
    return StreamingResponse(
        BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{inv.invoice_number}_signed.pdf"'},
    )


@router.get("/{invoice_id}/cost-of-sales/pdf")
def download_invoice_cost_of_sales_pdf(
    invoice_id: int,
    db: Session = Depends(get_db),
    _: models.User = Depends(get_current_user),
):
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    from utils.pdf_generator import generate_cost_of_sales_pdf

    inv = db.query(models.Invoice).filter(models.Invoice.id == invoice_id).first()
    if not inv:
        raise HTTPException(404, "Invoice not found")

    item_rows = (
        db.query(
            models.InvoiceItem.product_id,
            models.Product.product_name,
            func.sum(models.InvoiceItem.quantity).label("qty"),
            func.sum(models.InvoiceItem.cost_price * models.InvoiceItem.quantity).label("cost"),
            func.sum(models.InvoiceItem.line_total).label("revenue"),
        )
        .join(models.Product, models.Product.id == models.InvoiceItem.product_id)
        .filter(models.InvoiceItem.invoice_id == invoice_id)
        .group_by(models.InvoiceItem.product_id, models.Product.product_name)
        .order_by(func.sum(models.InvoiceItem.cost_price * models.InvoiceItem.quantity).desc())
        .all()
    )

    total_cost = sum(float(product_row.cost or 0) for product_row in item_rows)
    total_revenue = sum(float(product_row.revenue or 0) for product_row in item_rows)
    gross_profit = total_revenue - total_cost
    gross_margin = round(gross_profit / total_revenue * 100, 2) if total_revenue else 0

    report_data = {
        "meta": {
            "invoice_number": inv.invoice_number,
            "customer_name": inv.customer.customer_name if inv.customer else "",
        },
        "summary": {
            "total_cost": total_cost,
            "total_revenue": total_revenue,
            "gross_profit": gross_profit,
            "gross_margin_pct": gross_margin,
        },
        "by_product": [
            {
                "product_id": product_row.product_id,
                "product_name": product_row.product_name,
                "qty": float(product_row.qty or 0),
                "unit_cost_price": (float(product_row.cost or 0) / float(product_row.qty or 0)) if float(product_row.qty or 0) > 0 else 0,
                "cost": float(product_row.cost or 0),
                "revenue": float(product_row.revenue or 0),
                "gross_profit": float(product_row.revenue or 0) - float(product_row.cost or 0),
                "margin_pct": round(
                    (float(product_row.revenue or 0) - float(product_row.cost or 0)) / float(product_row.revenue) * 100, 2
                ) if product_row.revenue else 0,
            }
            for product_row in item_rows
        ],
        "by_invoice": [
            {
                "invoice_id": inv.id,
                "invoice_number": inv.invoice_number,
                "invoice_date": str(inv.invoice_date),
                "customer_name": inv.customer.customer_name if inv.customer else "",
                "cost": total_cost,
                "revenue": total_revenue,
                "gross_profit": gross_profit,
                "margin_pct": gross_margin,
            }
        ],
    }
    pdf_bytes = generate_cost_of_sales_pdf(
        report_data,
        title_suffix=f" ({inv.invoice_number})",
        cost_only=True,
    )
    return StreamingResponse(
        BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{inv.invoice_number}_cost_of_sales.pdf"'},
    )


@router.post("/{invoice_id}/upload-pdf", response_model=schemas.InvoiceOut)
async def upload_invoice_pdf(
    invoice_id: int,
    file: UploadFile = File(...),
    additional_emails: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    _: models.User = Depends(get_current_user),
):
    """Upload custom PDF, send to primary/additional recipients, and store for reuse."""
    import uuid
    from utils.s3 import upload_bytes, delete_object

    if not file.filename or not file.filename.lower().endswith(".pdf"):
        raise HTTPException(400, "Only PDF files are accepted")

    inv = db.query(models.Invoice).filter(models.Invoice.id == invoice_id).first()
    if not inv:
        raise HTTPException(404, "Invoice not found")

    # Delete the old custom PDF from S3 if it exists
    if inv.custom_pdf_s3_key:
        delete_object(inv.custom_pdf_s3_key)

    content = await file.read()
    s3_key = f"invoices/{invoice_id}/custom_{uuid.uuid4()}.pdf"
    upload_bytes(s3_key, content, "application/pdf")

    inv.custom_pdf_s3_key = s3_key
    db.commit()
    db.refresh(inv)
    recipients: List[str] = [INVOICE_PRIMARY_RECIPIENT]
    if additional_emails:
        recipients.extend([e.strip() for e in additional_emails.split(",") if e.strip()])
    recipients = list(dict.fromkeys(recipients))
    if recipients:
        task = send_invoice_to_recipients_task.delay(inv.id, recipients)
        log_queue_event(
            db,
            task_id=task.id,
            event_type="invoice_email",
            title=f"Send invoice email {inv.invoice_number}",
            requested_by=_.id if _ else None,
            metadata={"invoice_id": inv.id, "recipients": recipients},
        )
    return inv


@router.delete("/{invoice_id}/upload-pdf", response_model=schemas.InvoiceOut)
def remove_invoice_pdf(
    invoice_id: int,
    db: Session = Depends(get_db),
    _: models.User = Depends(get_current_user),
):
    """Remove the uploaded PDF so the system-generated one is used again."""
    from utils.s3 import delete_object

    inv = db.query(models.Invoice).filter(models.Invoice.id == invoice_id).first()
    if not inv:
        raise HTTPException(404, "Invoice not found")

    if inv.custom_pdf_s3_key:
        delete_object(inv.custom_pdf_s3_key)
        inv.custom_pdf_s3_key = None
        db.commit()
        db.refresh(inv)
    return inv


@router.post("/{invoice_id}/upload-signed", response_model=schemas.InvoiceOut)
async def upload_signed_invoice(
    invoice_id: int,
    file: UploadFile = File(...),
    additional_emails: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    """Upload signed invoice (PDF or image), mark completed, and store for future viewing."""
    import uuid
    from io import BytesIO
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.utils import ImageReader
    from reportlab.pdfgen import canvas
    from utils.s3 import upload_bytes, delete_object

    filename = (file.filename or "").lower()
    allowed_image_exts = (".jpg", ".jpeg", ".png", ".webp")
    is_pdf = filename.endswith(".pdf")
    is_image = filename.endswith(allowed_image_exts)
    if not filename or (not is_pdf and not is_image):
        raise HTTPException(400, "Only PDF or image files (JPG, PNG, WEBP) are accepted")

    inv = db.query(models.Invoice).filter(models.Invoice.id == invoice_id).first()
    if not inv:
        raise HTTPException(404, "Invoice not found")
    if inv.status == models.InvoiceStatus.cancelled:
        raise HTTPException(400, "Cancelled invoice cannot be completed")
    if not inv.quotation_id:
        raise HTTPException(400, "Signed upload is only available for converted invoices")

    if inv.custom_pdf_s3_key:
        delete_object(inv.custom_pdf_s3_key)

    content = await file.read()
    if is_image:
        image = ImageReader(BytesIO(content))
        img_w, img_h = image.getSize()
        page_w, page_h = A4
        scale = min(page_w / img_w, page_h / img_h)
        draw_w = img_w * scale
        draw_h = img_h * scale
        x = (page_w - draw_w) / 2
        y = (page_h - draw_h) / 2
        out = BytesIO()
        pdf = canvas.Canvas(out, pagesize=A4)
        pdf.drawImage(image, x, y, width=draw_w, height=draw_h, preserveAspectRatio=True, anchor="c")
        pdf.showPage()
        pdf.save()
        content = out.getvalue()
    s3_key = f"invoices/{invoice_id}/signed_{uuid.uuid4()}.pdf"
    upload_bytes(s3_key, content, "application/pdf")

    inv.custom_pdf_s3_key = s3_key
    inv.status = models.InvoiceStatus.completed
    audit.log(
        db,
        models.AuditAction.update,
        models.AuditEntity.invoice,
        inv.id,
        current_user.id if current_user else None,
        description=f"Uploaded signed invoice for {inv.invoice_number} and marked completed",
        new_values={"status": models.InvoiceStatus.completed.value, "signed_pdf_s3_key": s3_key},
    )
    db.commit()
    db.refresh(inv)

    # Do not block completion flow on Make integration issues.
    try:
        send_document_to_make_from_s3(
            doc_type="invoice",
            document_number=inv.invoice_number,
            s3_key=s3_key,
            filename=f"{inv.invoice_number}.pdf",
            customer_name=inv.customer.customer_name if inv.customer else "",
        )
    except Exception:
        pass

    return inv


@router.post("/{invoice_id}/generate-pdf", response_model=schemas.JobEnqueuedResponse,
             status_code=202)
def generate_invoice_pdf(
    invoice_id: int,
    db: Session = Depends(get_db),
    _: models.User = Depends(get_current_user),
):
    """
    Queue PDF generation. Returns a task_id immediately (< 5 ms).
    Poll GET /api/v1/jobs/{task_id}, then download via GET /api/v1/jobs/{task_id}/download.
    """
    inv = db.query(models.Invoice).filter(models.Invoice.id == invoice_id).first()
    if not inv:
        raise HTTPException(404, "Invoice not found")

    task = generate_invoice_pdf_task.delay(invoice_id)
    log_queue_event(
        db,
        task_id=task.id,
        event_type="invoice_pdf",
        title=f"Generate invoice PDF {inv.invoice_number}",
        requested_by=_.id if _ else None,
        metadata={"invoice_id": inv.id},
    )
    return schemas.JobEnqueuedResponse(
        task_id=task.id,
        message=f"PDF generation queued for {inv.invoice_number}. "
                f"Poll /api/v1/jobs/{task.id} for status.",
    )


@router.post("/{invoice_id}/send-email", response_model=schemas.MessageResponse)
def send_invoice_email(
    invoice_id: int,
    body: InvoiceSendEmailRequest,
    db: Session = Depends(get_db),
    _: models.User = Depends(get_current_user),
):
    """Send invoice PDF to the customer's email and/or additional email addresses."""
    inv = db.query(models.Invoice).filter(models.Invoice.id == invoice_id).first()
    if not inv:
        raise HTTPException(404, "Invoice not found")

    emails: List[str] = []
    if inv.customer and inv.customer.email:
        emails.append(inv.customer.email)
    emails.append(INVOICE_PRIMARY_RECIPIENT)
    if body.additional_emails:
        emails.extend([e.strip() for e in body.additional_emails if e.strip()])
    emails = list(dict.fromkeys(emails))

    if not emails:
        raise HTTPException(400, "No email addresses to send to")

    task = send_invoice_to_recipients_task.delay(inv.id, emails)
    log_queue_event(
        db,
        task_id=task.id,
        event_type="invoice_email",
        title=f"Send invoice email {inv.invoice_number}",
        requested_by=_.id if _ else None,
        metadata={"invoice_id": inv.id, "recipients": emails},
    )
    return schemas.MessageResponse(
        message=f"Invoice email queued for {len(emails)} recipient(s)"
    )


@router.post("/{invoice_id}/upload-to-make", response_model=schemas.MessageResponse)
def upload_invoice_to_make(
    invoice_id: int,
    body: InvoiceSendEmailRequest,
    db: Session = Depends(get_db),
    _: models.User = Depends(get_current_user),
):
    """Send the existing invoice PDF to primary and additional recipients."""
    inv = db.query(models.Invoice).filter(models.Invoice.id == invoice_id).first()
    if not inv:
        raise HTTPException(404, "Invoice not found")

    if inv.custom_pdf_s3_key:
        send_document_to_make_from_s3(
            doc_type="invoice",
            document_number=inv.invoice_number,
            s3_key=inv.custom_pdf_s3_key,
            filename=f"{inv.invoice_number}.pdf",
            customer_name=inv.customer.customer_name if inv.customer else "",
        )

    emails: List[str] = [INVOICE_PRIMARY_RECIPIENT]
    if body.additional_emails:
        emails.extend([e.strip() for e in body.additional_emails if e.strip()])
    emails = list(dict.fromkeys(emails))
    if not emails:
        raise HTTPException(400, "No email addresses to send to")

    task = send_invoice_to_recipients_task.delay(inv.id, emails)
    log_queue_event(
        db,
        task_id=task.id,
        event_type="invoice_upload_to_make",
        title=f"Upload invoice to make {inv.invoice_number}",
        requested_by=_.id if _ else None,
        metadata={"invoice_id": inv.id, "recipients": emails},
    )
    return schemas.MessageResponse(
        message=f"Invoice upload-to-make queued for {len(emails)} recipient(s)"
    )


@router.post("/{invoice_id}/cancel", response_model=schemas.InvoiceOut)
def cancel_invoice(
    invoice_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin_or_manager),
):
    inv = db.query(models.Invoice).filter(models.Invoice.id == invoice_id).first()
    if not inv:
        raise HTTPException(404, "Invoice not found")
    if inv.status == models.InvoiceStatus.cancelled:
        raise HTTPException(400, "Invoice is already cancelled")
    if inv.status == models.InvoiceStatus.completed:
        raise HTTPException(400, "Completed invoice cannot be cancelled")
    inv.status = models.InvoiceStatus.cancelled
    audit.log(db, models.AuditAction.cancel, models.AuditEntity.invoice, inv.id,
               current_user.id, description=f"Cancelled invoice {inv.invoice_number}")
    db.commit()
    db.refresh(inv)
    return inv


@router.delete("/{invoice_id}", response_model=schemas.MessageResponse)
def delete_invoice(
    invoice_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
):
    msg = _delete_invoice_record(db, invoice_id, current_user)
    db.commit()
    return msg

